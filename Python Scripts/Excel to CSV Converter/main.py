"""Excel to CSV Converter — CLI tool.

Convert .xlsx / .xls files to CSV.
Supports multi-sheet workbooks (one CSV per sheet or merged).
Uses openpyxl if available; falls back to csv-only mode for .csv inputs.

Usage:
    python main.py
    python main.py file.xlsx
    python main.py file.xlsx --sheet Sheet1 --output out.csv
"""

import argparse
import csv
import sys
from pathlib import Path


def excel_to_csv_openpyxl(xlsx_path: Path, output_dir: Path,
                            sheet_name: str | None = None) -> list[Path]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames
    outputs = []
    for name in sheets:
        ws  = wb[name]
        safe_name = name.replace("/", "_").replace("\\", "_")
        out = output_dir / f"{xlsx_path.stem}_{safe_name}.csv"
        with open(out, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([("" if v is None else str(v)) for v in row])
        outputs.append(out)
        print(f"  Sheet '{name}' → {out}  ({ws.max_row} rows, {ws.max_column} cols)")
    wb.close()
    return outputs


def excel_to_csv_xlrd(xls_path: Path, output_dir: Path,
                       sheet_name: str | None = None) -> list[Path]:
    import xlrd
    wb = xlrd.open_workbook(str(xls_path))
    names = [sheet_name] if sheet_name and sheet_name in wb.sheet_names() else wb.sheet_names()
    outputs = []
    for name in names:
        ws  = wb.sheet_by_name(name)
        safe = name.replace("/","_")
        out  = output_dir / f"{xls_path.stem}_{safe}.csv"
        with open(out, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for rx in range(ws.nrows):
                writer.writerow([str(ws.cell_value(rx, cx)) for cx in range(ws.ncols)])
        outputs.append(out)
        print(f"  Sheet '{name}' → {out}  ({ws.nrows} rows, {ws.ncols} cols)")
    return outputs


def convert(path: Path, output_dir: Path, sheet: str | None = None) -> list[Path]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        try:
            return excel_to_csv_openpyxl(path, output_dir, sheet)
        except ImportError:
            print("  openpyxl not installed. Run: pip install openpyxl")
            return []
    elif suffix == ".xls":
        try:
            return excel_to_csv_xlrd(path, output_dir, sheet)
        except ImportError:
            print("  xlrd not installed. Run: pip install xlrd")
            return []
    elif suffix == ".csv":
        print(f"  {path.name} is already a CSV file.")
        return [path]
    else:
        print(f"  Unsupported format: {suffix}")
        return []


def list_sheets(path: Path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        print(f"  Sheets in {path.name}:")
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"    • {name}  ({ws.max_row} rows)")
        wb.close()
    except ImportError:
        print("  openpyxl not installed.")


def main():
    parser = argparse.ArgumentParser(description="Excel to CSV Converter")
    parser.add_argument("file",    nargs="?")
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--output","-o", default=None)
    parser.add_argument("--list-sheets", action="store_true")
    args = parser.parse_args()

    if args.file:
        path = Path(args.file)
        if not path.exists():
            print(f"File not found: {path}")
            sys.exit(1)
        if args.list_sheets:
            list_sheets(path)
            return
        out_dir = Path(args.output) if args.output else path.parent
        out_dir.mkdir(parents=True, exist_ok=True)
        convert(path, out_dir, args.sheet)
        return

    print("Excel to CSV Converter")
    print("────────────────────────────────")

    while True:
        raw = input("\nExcel file path (or 'q'): ").strip()
        if raw.lower() == "q":
            break
        path = Path(raw)
        if not path.exists():
            print(f"  Not found: {path}")
            continue

        if path.suffix.lower() in (".xlsx", ".xls"):
            list_sheets(path)
            sheet = input("  Sheet name (blank = all sheets): ").strip() or None
        else:
            sheet = None

        out_dir = path.parent
        results = convert(path, out_dir, sheet)
        if results:
            print(f"\n  Converted {len(results)} sheet(s) successfully.")


if __name__ == "__main__":
    main()
